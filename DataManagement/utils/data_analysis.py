import pandas as pd
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
import io
import base64
import matplotlib
matplotlib.use('Agg')
from PyPDF2 import PdfReader
import re
from collections import Counter
from wordcloud import WordCloud  # pip install wordcloud
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
from openpyxl import load_workbook 
import pdfplumber
from datetime import datetime

def analyze_data(file_path, operation, parameters):
      # Determine file type
    if file_path.endswith(('.csv', '.xlsx', '.xls', '.pdf')):
        return analyze_tabular_data(file_path, operation, parameters)
    else:
        raise ValueError("Unsupported file format")

def analyze_tabular_data(file_path, operation, parameters):
    """Enhanced tabular data analysis function with improved PDF and Excel handling"""
    
    # Read file based on extension
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
    elif file_path.endswith('.pdf'):
        # Process PDF with pdfplumber for better table extraction
        tables = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables():
                    # Convert each table to a DataFrame
                    df_table = pd.DataFrame(table[1:], columns=table[0])
                    tables.append(df_table)
        
        if not tables:
            raise ValueError("No tables found in PDF")
        
        # For simplicity, we'll use the first table found
        # In a production environment, you might want to return all tables
        df = tables[0]
        
    elif file_path.endswith(('.xlsx', '.xls')):
        # Get all sheet names first
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        # Filter out empty sheets
        non_empty_sheets = []
        for sheet in sheet_names:
            sheet_df = pd.read_excel(file_path, sheet_name=sheet)
            if not sheet_df.empty:
                non_empty_sheets.append({
                    'sheet_name': sheet,
                    'dataframe': sheet_df
                })
        
        if not non_empty_sheets:
            raise ValueError("No data found in any Excel sheet")
        
        # For operations that can work with multiple sheets, we might process all
        # For now, we'll use the first non-empty sheet for compatibility
        df = non_empty_sheets[0]['dataframe']
        
        # Store sheet info in parameters for potential multi-sheet operations
        parameters['_excel_sheets'] = non_empty_sheets
    else:
        raise ValueError("Unsupported file format")
    
    # Clean the dataframe by converting numeric columns
    for col in df.columns:
        # First try to convert to datetime
        try:
            df[col] = pd.to_datetime(df[col], errors='raise', infer_datetime_format=True)
            continue  # Skip numeric conversion if it's a date
        except (ValueError, TypeError):
            pass
        
        # Then try to convert to numeric
        try:
            # Try converting to float first
            numeric_vals = pd.to_numeric(df[col], errors='raise')
            
            # Check if it's actually integer (all whole numbers)
            if all(x == int(x) for x in numeric_vals.dropna() if not pd.isna(x)):
                df[col] = numeric_vals.astype('Int64')  # Using Int64 to handle NaN
            else:
                df[col] = numeric_vals
        except (ValueError, TypeError):
            # If both conversions fail, leave as string
            pass
    
    # Perform requested operation
    if operation == 'describe':
        # For datetime columns, calculate different stats
        description = df.describe(include='all', datetime_is_numeric=True)
        
        # Convert results to JSON-compatible format
        result = {}
        for col in description.columns:
            col_stats = {}
            for stat, value in description[col].items():
                if pd.isna(value) or (isinstance(value, (float, int)) and np.isinf(value)):
                    col_stats[stat] = None
                elif isinstance(value, (pd.Timestamp, datetime)):
                    col_stats[stat] = value.isoformat()
                elif isinstance(value, (np.integer, np.floating)):
                    col_stats[stat] = int(value) if isinstance(value, np.integer) else float(value)
                else:
                    col_stats[stat] = str(value)
            result[col] = col_stats
        
        return {'result': result}
    
    elif operation == 'correlation':
        df = df.apply(pd.to_numeric, errors='coerce')  # Convert all columns to numeric
        df = df.dropna()  # Drop NaNs to avoid issues
        correlation_matrix = df.corr()
        
        # Convert numpy types to native Python types for JSON serialization
        result = {
            col: {
                other_col: float(value) if pd.notna(value) else None
                for other_col, value in row.items()
            }
            for col, row in correlation_matrix.to_dict().items()
        }
        
        return {'result': result}
    
    elif operation == 'value_counts':
        column = parameters.get('column')
        if not column or column not in df.columns:
            raise ValueError(f"Column '{column}' not found in data")
        
        counts = df[column].value_counts().to_dict()
        
        # Generate and save plot to a bytes buffer
        buf = io.BytesIO()
        plt.figure(figsize=(10, 6))
        df[column].value_counts().plot(kind='barh')
        plt.title(f'Value Counts for {column}')
        plt.xlabel('Count')
        plt.ylabel('Value')
        plt.tight_layout()
        plt.savefig(buf, format='png')
        plt.close()
        
        # Convert to base64 for easy transfer
        buf.seek(0)
        image_base64 = base64.b64encode(buf.read()).decode('utf-8')
        
        return {
            'result': counts,
            'plot': image_base64
        }
    
    elif operation == 'linear_regression':
        x_col = parameters.get('x_column')
        y_col = parameters.get('y_column')
        
        if not x_col or not y_col or x_col not in df.columns or y_col not in df.columns:
            raise ValueError("Both x_column and y_column must be valid columns")
        
        # Convert columns to numeric, forcing errors to NaN
        df[x_col] = pd.to_numeric(df[x_col], errors='coerce')
        df[y_col] = pd.to_numeric(df[y_col], errors='coerce')
        
        # Drop NaN values
        df = df.dropna(subset=[x_col, y_col])
        
        # Ensure we have at least two valid data points
        if len(df) < 2:
            raise ValueError("Not enough valid data points for linear regression")
        
        # Ensure no infinite values
        if np.isinf(df[x_col]).any() or np.isinf(df[y_col]).any():
            raise ValueError("Columns contain infinite values")
        
        x = df[x_col].values
        y = df[y_col].values
        
        slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
        
        # Generate plot
        plt.figure(figsize=(10, 6))
        plt.scatter(x, y, label='Data points')
        plt.plot(x, slope * x + intercept, color='red', label='Regression line')
        plt.xlabel(x_col)
        plt.ylabel(y_col)
        plt.title(f'Linear Regression: {y_col} vs {x_col}')
        plt.legend()
        
        # Save plot to base64
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plot_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close()
        
        result = {
            'slope': slope,
            'intercept': intercept,
            'r_squared': r_value ** 2,
            'p_value': p_value,
            'std_err': std_err,
            'plot': plot_base64
        }
    
    elif operation == 'cluster':
        n_clusters = parameters.get('n_clusters', 3)
        features = parameters.get('features', [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])])
        
        if not features:
            raise ValueError("No numeric features available for clustering")
        
        df[features] = df[features].apply(pd.to_numeric, errors='coerce')
        df = df.dropna()
        
        if df.shape[0] < n_clusters:
            raise ValueError("Not enough valid data points for clustering")
        
        X = df[features].values
        X = (X - X.mean(axis=0)) / X.std(axis=0)
        
        kmeans = KMeans(n_clusters=n_clusters, random_state=42)
        clusters = kmeans.fit_predict(X)
        df['cluster'] = clusters
        
        plt.figure(figsize=(10, 6))
        plt.scatter(X[:, 0], X[:, 1], c=clusters, cmap='viridis')
        plt.xlabel(features[0])
        plt.ylabel(features[1])
        plt.title(f'K-Means Clustering (k={n_clusters})')
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plot_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close()
        
        result = {
            'cluster_centers': kmeans.cluster_centers_.tolist(),
            'inertia': kmeans.inertia_,
            'plot': plot_base64,
            'sample_clusters': df[[features[0], features[1], 'cluster']].head(10).to_dict('records')
        }
    
    elif operation == 'pca':
        n_components = parameters.get('n_components', 2)
        features = parameters.get('features', [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])])
        
        if not features:
            raise ValueError("No numeric features available for PCA")
        
        df[features] = df[features].apply(pd.to_numeric, errors='coerce')
        df = df.dropna()
        
        if df.shape[0] < n_components:
            raise ValueError("Not enough valid data points for PCA")
        
        X = df[features].values
        X = (X - X.mean(axis=0)) / X.std(axis=0)
        
        pca = PCA(n_components=n_components)
        components = pca.fit_transform(X)
        
        plt.figure(figsize=(10, 6))
        plt.scatter(components[:, 0], components[:, 1])
        plt.xlabel('Principal Component 1')
        plt.ylabel('Principal Component 2')
        plt.title('PCA Results')
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plot_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close()
        
        result = {
            'explained_variance_ratio': pca.explained_variance_ratio_.tolist(),
            'components': pca.components_.tolist(),
            'plot': plot_base64
        }
    else:
        raise ValueError(f"Unsupported operation: {operation}")
    
    return {
        'operation': operation,
        'parameters': parameters,
        'result': result
    }