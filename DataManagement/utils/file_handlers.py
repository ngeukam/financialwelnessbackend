import pandas as pd
from PyPDF2 import PdfReader
import csv
import os
import pdfplumber
import numpy as np
from typing import List, Dict, Union, Optional
from datetime import datetime
import dateutil.parser
import re
from openpyxl import load_workbook
import difflib

def handle_uploaded_file(file, index: Optional[Union[int, List[int]]] = None, columns_to_remove: Optional[Union[str, List[str]]] = None, min_empty_values: Optional[int] = None) -> Dict[str, any]:
    """
    Unified file upload handler that returns consistent structure for all file types
    
    """
    file_type = file.name.split('.')[-1].lower()
    
    try:
        if file_type == 'pdf':
            data = process_pdf(file, index, columns_to_remove, min_empty_values)
        elif file_type in ['xlsx', 'xls']:
            data = process_tabular(file, index, columns_to_remove, min_empty_values)
        elif file_type == 'csv':
            data = process_tabular(file, index, columns_to_remove, min_empty_values)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        return {
            'success': True,
            'data': data,
            'file_type': file_type
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'file_type': file_type,
            'data': {
                'status': 'error',
                'message': str(e),
                'total_tables': 0,
                'multi_page_tables': 0,
                'tables': []
            }
        }
    
def process_pdf(file, index: Optional[Union[int, List[int]]] = None, columns_to_remove: Optional[Union[str, List[str]]] = None, min_empty_values: Optional[int] = None) -> Dict[str, Union[List[Dict], int, str]]:
    """
    Extract tables from PDF, combining tables that span multiple pages.
    """
    try:
        # First pass: Extract all raw tables with their positions
        raw_tables = []
        with pdfplumber.open(file) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables({
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines"
                })
                
                for table_num, table_data in enumerate(tables, start=1):
                    if not table_data or len(table_data) < 1:
                        continue
                    
                    # Get table bounding box
                    table_bbox = page.find_tables()[table_num-1].bbox
                    
                    raw_tables.append({
                        'page': page_num,
                        'table_num': table_num,
                        'data': table_data,
                        'bbox': table_bbox  # (x0, top, x1, bottom)
                    })
        
        if not raw_tables:
            return {
                'status': 'error',
                'message': 'No tables found in PDF',
                'total_tables': 0,
                'multi_page_tables': 0
            }
        
        # Second pass: Identify and combine multi-page tables
        processed_tables = []
        current_table = None
        
        for i, table in enumerate(raw_tables):
            if current_table is None:
                # Start new table
                current_table = {
                    'table_id': len(processed_tables) + 1,
                    'headers': [str(h).strip() for h in table['data'][0]],
                    'data_rows': table['data'][1:] if len(table['data']) > 1 else [],
                    'pages': [table['page']],
                    'bottom_pos': table['bbox'][3]  # y-position of table bottom
                }
            else:
                # Check if this is a continuation (table at top of new page)
                is_continuation = (
                    table['page'] == current_table['pages'][-1] + 1 and
                    table['bbox'][1] < 100  # Table starts near top of page
                )
                
                if is_continuation:
                    # Continue the current table
                    current_table['data_rows'].extend(table['data'])
                    current_table['pages'].append(table['page'])
                    current_table['bottom_pos'] = table['bbox'][3]
                else:
                    # Finalize current table and start new one
                    processed_tables.append(finalize_table(current_table, index, columns_to_remove, min_empty_values))
                    current_table = {
                        'table_id': len(processed_tables) + 1,
                        'headers': [str(h).strip() for h in table['data'][0]],
                        'data_rows': table['data'][1:] if len(table['data']) > 1 else [],
                        'pages': [table['page']],
                        'bottom_pos': table['bbox'][3]
                    }
        
        # Add the last table being processed
        if current_table is not None:
            processed_tables.append(finalize_table(current_table, index, columns_to_remove, min_empty_values))
        
        # Count multi-page tables
        multi_page_count = sum(1 for t in processed_tables if len(t['pages']) > 1)
        
        return {
            'status': 'success',
            'tables': processed_tables,
            'total_tables': len(processed_tables),
            'multi_page_tables': multi_page_count,
            'message': f'Extracted {len(processed_tables)} tables ({multi_page_count} multi-page)'
        }
        
    except Exception as e:
        return {
            'status': 'error',
            'message': f'Error processing PDF: {str(e)}',
            'total_tables': 0,
            'multi_page_tables': 0
        }

def finalize_table(table: Dict, index: Optional[Union[int, List[int]]] = None, columns_to_remove: Optional[Union[str, List[str]]] = None, min_empty_values: Optional[int] = None) -> Dict:
    """Convert raw table data to cleaned final format"""
    df = pd.DataFrame(table['data_rows'], columns=table['headers'])
    df = clean_dataframe(df, min_empty_values=min_empty_values)
    if index is not None:
        indices = [index] if isinstance(index, int) else index
        for idx in indices:
            if abs(int(idx)) >= len(df):  # Handle negative indices
                raise ValueError(f"Index {idx} out of bounds for table with {len(df)} rows")
        df = remove_row_by_index(df, indices)
    
    # Remove specified columns
    if columns_to_remove:
        df = remove_column(df, columns_to_remove)
        # Update headers to match remaining columns
        table['headers'] = list(df.columns)
    return {
        'table_id': table['table_id'],
        'headers': table['headers'],
        'data': df.replace({np.nan: None}).to_dict('records'),
        'shape': df.shape,
        'pages': table['pages'],
        'is_multi_page': len(table['pages']) > 1
    }

# Fonction pour vérifier si une colonne contient majoritairement des valeurs numériques sous forme de chaînes
def is_numeric_string_column(series):
    """
    Vérifie si une colonne contient majoritairement des valeurs numériques sous forme de chaînes.
    Prend en compte les espaces, séparateurs de milliers et décimales.
    """
    if series.empty:
        return False
    
    # On garde seulement les valeurs non-nulles pour le test
    non_null = series.dropna()
    
    # Vérifie le pattern numérique avec possibilité de :
    # - espaces/séparateurs
    # - signe +/-
    # - décimales (virgule ou point)
    pattern = r'^[+-]?[\d\s\u00A0]*[,.]?\d+$'
    # pattern = r'^[+-]?[\d\s\u00A0.,]*[,.]?\d+\s*[^\d.,-+]*$'
    
    # Au moins 75% des valeurs doivent matcher le pattern numérique
    return non_null.str.contains(pattern, regex=True).mean() > 0.75

def is_date(value):
    if not value or pd.isna(value):  # Vérifier si la valeur est None ou NaN
        return None
    if re.search(r'[a-zA-Z]', value) and re.search(r'\d', value):  
        return None  # Contient chiffres + lettres => Probablement pas une date valide
    try:
        parsed_date = dateutil.parser.parse(value, dayfirst=True)
        if 1900 <= parsed_date.year <= 2050:
            return parsed_date
    except (ValueError, TypeError):
        return None
    return None

def clean_dataframe(df: pd.DataFrame, min_empty_values: Optional[int] = None) -> pd.DataFrame:

    pd.set_option('future.no_silent_downcasting', True)

    df = df.copy()
    # Formats de date à essayer (par ordre de priorité)
    DATE_FORMATS = [
        '%Y-%m-%d',   # ISO format (2023-12-31)
        '%d/%m/%Y',   # Français (31/12/2023)
        '%m/%d/%Y',   # US (12/31/2023)
        '%d-%m-%Y',   # Français avec tirets (31-12-2023)
        '%Y%m%d',     # Compact (20231231)
        '%d %b %Y',   # 31 Dec 2023
        '%d %B %Y',   # 31 December 2023
    ]
    
    for col in df.columns:
        # Nettoyage de base : remplacement des valeurs vides (sans strip pour garder les espaces)
        col_data = df[col].astype(str)
        # col_data = col_data.replace(["None", "null", "", "nan", "NaN", "NAN", "NA"], np.nan)
        col_data = df[col].astype(str).replace(["None", "null", "", "nan", "NaN", "NAN", "NA"], None)

        date_converted = False
        for date_format in DATE_FORMATS:
            try:
                dates = pd.to_datetime(col_data, format=date_format, errors='coerce')
                if dates.notna().any():  # Si au moins une date valide est trouvée
                    df[col] = dates.dt.strftime("%Y-%m-%d")
                    date_converted = True
                    break
            except:
                continue
            
        df[col] = df[col].where(pd.notna(df[col]), None)
        if date_converted:
            continue  # On passe à la colonne suivante si conversion réussie
        
        # Détection des colonnes numériques avec votre fonction
        if is_numeric_string_column(col_data.dropna()):
            # Création d'une copie pour la conversion numérique
            num_data = col_data.copy()
            
            # Nettoyage spécifique pour la conversion numérique
            num_data = num_data.str.replace(r'[\s\u00A0]', '', regex=True)  # Supprime espaces
            num_data = num_data.str.replace(',', '.', regex=False)          # Remplace virgule par point
            
            # Conversion en numérique
            numeric_vals = pd.to_numeric(num_data, errors="coerce")
            
            if numeric_vals.notna().any():
                # Conversion en Int64 si entier, sinon float
                if (numeric_vals.dropna() % 1 == 0).all():
                    df[col] = numeric_vals.astype('Int64')
                else:
                    df[col] = numeric_vals

    # Filtrage des lignes avec trop de NaN (si demandé)
    if min_empty_values is not None:
        nan_counts = df.isna().sum(axis=1)
        df = df[nan_counts < min_empty_values].copy()
    return df


def remove_row_by_index(df, index):
    """
    Supprime une ligne du DataFrame en fonction de l'index donné.

    Args:
        df: DataFrame pandas
        index: Index de la ligne à supprimer (peut être un seul index ou une liste d'index)
    
    Returns:
        DataFrame sans la/les ligne(s) supprimée(s)
    
    Exemples:
        >>> df = remove_row_by_index(df, 5)          # Supprime la ligne d'index 5
        >>> df = remove_row_by_index(df, [1, 3, 5])  # Supprime les lignes 1, 3 et 5
    """
    try:
        if isinstance(index, (list, tuple, set)):
            # Suppression multiple
            existing_indices = [idx for idx in index if idx in df.index]
            if existing_indices:
                df = df.drop(existing_indices)
                print(f"Rows {existing_indices} deleted successfully!")
            else:
                print("None of the specified indices exist in the DataFrame.")
        else:
            # Suppression unique
            if index in df.index:
                df = df.drop(index)
                print(f"Row {index} deleted successfully!")
            else:
                print(f"Row {index} does not exist in the DataFrame.")
    
    except Exception as e:
        print(f"Error while deleting rows: {str(e)}")
    
    return df

def remove_column(df, column_name):
    try:
        if isinstance(column_name, (list, tuple, set)):
            # Vérifier chaque colonne et proposer une correction si nécessaire
            corrected_columns = []
            for col in column_name:
                if col in df.columns:
                    corrected_columns.append(col)
                else:
                    closest_match = difflib.get_close_matches(col, df.columns, n=1, cutoff=0.6)
                    if closest_match:
                        print(f"⚠️ Column '{col}' not found. Did you mean '{closest_match[0]}'?")
                        corrected_columns.append(closest_match[0])  # Utiliser la meilleure suggestion

            if corrected_columns:
                df = df.drop(columns=corrected_columns)
                print(f"Columns {corrected_columns} deleted successfully!")
            else:
                print("None of the specified columns exist in the DataFrame.")
        
        else:
            # Vérifier une seule colonne
            if column_name in df.columns:
                df = df.drop(columns=[column_name])
                print(f"Column '{column_name}' deleted successfully!")
            else:
                closest_match = difflib.get_close_matches(column_name, df.columns, n=1, cutoff=0.6)
                if closest_match:
                    print(f"⚠️ Column '{column_name}' not found. Did you mean '{closest_match[0]}'?")
                    df = df.drop(columns=[closest_match[0]])
                else:
                    print(f"Column '{column_name}' does not exist in the DataFrame.")
        print('df', df)
        return df  # Retourner le DataFrame modifié

    except Exception as e:
        print(f"Error while deleting columns: {str(e)}")
        return df  # Retourner le DataFrame original en cas d'erreur



def process_tabular(file, index: Optional[Union[int, List[int]]] = None, columns_to_remove: Optional[Union[str, List[str]]] = None, min_empty_values: Optional[int] = None) -> Dict[str, Union[List[Dict], int, str]]:
    """Process CSV/Excel files to match PDF output structure"""
    try:
        # Read file based on extension
        if file.name.endswith('.csv'):
            df = pd.read_csv(file, thousands=',')
        else:  # Excel
            # Convert Excel file to table structure first
            try:
                wb = load_workbook(filename=file, read_only=True, data_only=True)
                sheets_data = {}
                
                # Use sheetnames instead of sheet_names (OpenPyXL compatibility)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    data = []
                    # Determine max column based on non-empty cells
                    max_col = 0
                    for row in ws.iter_rows(values_only=True):
                        if row:
                            max_col = max(max_col, len(row))
                    
                    # Read all rows
                    for row in ws.iter_rows(values_only=True):
                        if row:
                            # Pad row with None if shorter than max_col
                            padded_row = list(row) + [None] * (max_col - len(row))
                            data.append(padded_row)
                        else:
                            data.append([None] * max_col)
                    
                    # Convert to DataFrame if data exists
                    if data:
                        # Use first non-empty row as headers
                        headers = next((row for row in data if any(cell is not None for cell in row)), None)
                        if headers:
                            header_idx = data.index(headers)
                            # Create DataFrame with proper headers
                            df = pd.DataFrame(data[header_idx+1:], columns=headers)
                            sheets_data[sheet_name] = df
                
                # For backward compatibility, use first sheet if multiple exist
                df = list(sheets_data.values())[0] if sheets_data else pd.DataFrame()

            except Exception as e:
                # Fallback to pandas read_excel if openpyxl fails
                df = pd.read_excel(file, thousands=',')


        if index is not None:
            indices = [index] if isinstance(index, int) else index
            for idx in indices:
                if abs(idx) >= len(df):  # Handle negative indices
                    raise ValueError(f"Index {idx} out of bounds for table with {len(df)} rows")
            df = remove_row_by_index(df, indices)
    
        # Remove specified columns
        if columns_to_remove :
            df = remove_column(df, columns_to_remove)
                
        # Clean dataframe
        df = clean_dataframe(df, min_empty_values)
        # Convert to standardized table format
        table = {
            'table_id': 1,
            'headers': [str(col) for col in df.columns],
            'data': df.to_dict('records'),
            'pages': [1],
            'is_multi_page': False
        }
        # Update headers to match remaining columns
        table['headers'] = list(df.columns)
        
        return {
            'status': 'success',
            'tables': [table],
            'total_tables': 1,
            'multi_page_tables': 0,
            'message': 'Successfully processed tabular file'
        }
        
    except Exception as e:
        return {
            'status': 'error',
            'message': f'Error processing tabular file: {str(e)}',
            'total_tables': 0,
            'multi_page_tables': 0
        }